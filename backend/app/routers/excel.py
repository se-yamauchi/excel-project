from fastapi import APIRouter, HTTPException, Form, File, UploadFile
from fastapi.responses import StreamingResponse
from io import BytesIO
import zipfile
import openpyxl

router = APIRouter()

# ダウンロードボタン押下時のファイル変換処理
@router.post("/upload-excel-folder")
async def upload_excel_folder(files: list[UploadFile] = File(...), magnification: int = Form(...)):
    try:
        output = BytesIO()
        with zipfile.ZipFile(output, 'w') as zip_file:
            for file in files:
                contents = await file.read()
                workbook = openpyxl.load_workbook(BytesIO(contents))
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet.views.sheetView[0].selection[0].activeCell = "A1"
                    sheet.views.sheetView[0].selection[0].sqref = "A1"
                    sheet.sheet_view.zoomScale = magnification

                for sheet_name in workbook.sheetnames:
                    workbook[sheet_name].views.sheetView[0].tabSelected = False
                first_sheet = workbook.sheetnames[0]
                workbook.active = workbook[first_sheet]

                modified_file = BytesIO()
                workbook.save(modified_file)
                modified_file.seek(0)
                zip_file.writestr(file.filename, modified_file.read())

        output.seek(0)
        return StreamingResponse(output, media_type='application/zip', headers={"Content-Disposition": "attachment; filename=changed_files.zip"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
